"""
Xero Processes — Munros Client
================================
Single public entry point: run_xpm_steps()

Called by the consumer for each family member when PMS = "XPM".
Currently downloads all Xero reports and saves them to the member output folder.
XPM client lookup and data update steps will be added here as the process map is confirmed.

iaa_rpa_xero_blue package (../iaa_rpa_xero_blue/src) must be on PYTHONPATH.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from iaa_rpa_framework import Config

from RPA_process.Client.Munros.workpaper_utility import (
    build_member_folder_path,
    ensure_folder,
    report_filename as _resolve_report_filename,
    financial_year_dates,
)
from RPA_process.Client.Munros import munros_constants as MC

from iaa_rpa_xero_blue.xero_blue_download_aged_paybles_summary_reports import (
    xero_blue_download_aged_payables_summary_reports,
)
from iaa_rpa_xero_blue.xero_blue_download_gst_reconcilliation_report import (
    xero_blue_download_gst_recconciliation_report,
)
from iaa_rpa_xero_blue.xero_blue_download_bank_reconciliation_report import (
    xero_blue_download_bank_reconciliation_report,
)
from iaa_rpa_xero_blue.xero_blue_download_payroll_employee_summary_report import (
    xero_blue_download_payroll_employee_summary_report,
)
from iaa_rpa_xero_blue.xero_blue_download_general_ledger_detail_report import (
    xero_blue_download_general_ledger_details_report,
)
from iaa_rpa_xero_blue.xero_blue_download_aged_receivable_summary_report import (
    xero_blue_download_aged_receivables_summary_report,
)

logger = logging.getLogger("IARPA." + __name__)

# Local alias so the extension string isn't repeated throughout every _dl_* function.
_EXTENSION = MC.REPORT_EXTENSION


# ---------------------------------------------------------------------------
# Private report download helpers
# ---------------------------------------------------------------------------


def _dl_balance_sheet(
    browser,
    client_name: str,
    fy: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement once xero_blue_download_balance_sheet_report exists in iaa_rpa_xero_blue."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [XPM] Balance Sheet — TODO: not yet in iaa_rpa_xero_blue (client=%s, fy=%s, file=%s)",
        client_name,
        fy,
        output_path.name,
    )
    _ = browser  # will be used once the download function is available
    return None


def _dl_profit_and_loss(
    browser,
    client_name: str,
    fy: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement once xero_blue_download_profit_and_loss_report exists in iaa_rpa_xero_blue."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [XPM] Profit and Loss — TODO: not yet in iaa_rpa_xero_blue (client=%s, fy=%s, file=%s)",
        client_name,
        fy,
        output_path.name,
    )
    _ = browser  # will be used once the download function is available
    return None


def _merge_bs_and_pl(
    bs_path: Optional[Path],
    pl_path: Optional[Path],
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """Merge Balance Sheet and P&L into one workbook with two named sheets."""
    if not bs_path or not pl_path:
        logger.info(
            "    [XPM] Merged financials skipped — BS and/or P&L not yet downloaded (TODO)"
        )
        return None

    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        merged_wb = openpyxl.Workbook()
        merged_wb.remove(merged_wb.active)
        # Sheet names are fixed by the workpaper standard — not config-driven.
        for source_path, sheet_title in (
            (bs_path, "Balance Sheet"),
            (pl_path, "Profit and Loss"),
        ):
            source_wb = openpyxl.load_workbook(source_path, data_only=True)
            dest_ws = merged_wb.create_sheet(title=sheet_title)
            for row in source_wb.active.iter_rows(values_only=True):
                dest_ws.append(list(row))
            source_wb.close()
        merged_wb.save(output_path)
        merged_wb.close()
        logger.info("    [XPM] Merged financials saved: %s", output_path.name)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Merged financials failed: %s", exc)
        return None


def _dl_aged_payables_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_aged_payables_summary_reports(
            browser=browser,
            client_name=client_name,
            xero_end_date=end_date,
            xero_financial_year=fy,
            is_add_gst_column=Config.get(MC.CFG_XERO_ADD_GST_COLUMN, "false").lower()
            == "true",
            xero_aging_by=Config.get(MC.CFG_XERO_AGING_BY, MC.DEFAULT_AGING_BY),
            window_title=window_title,
            download_directory=str(folder),
            report_file_name=filename,
            extension=_EXTENSION,
        )
        logger.info("    [XPM] Aged Payables Summary → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Aged Payables Summary failed: %s", exc)
        return None


def _dl_gst_reconciliation(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_gst_recconciliation_report(
            browser=browser,
            xero_client_name=client_name,
            xero_end_date=end_date,
            xero_financial_year=fy,
            xero_start_date=start_date,
            window_title=window_title,
            download_directory=str(folder),
            report_file_name=filename,
            xero_report_name=Config.get(
                MC.CFG_XERO_GST_REPORT_NAME, MC.XERO_RPT_GST_RECON
            ),
            extension=_EXTENSION,
        )
        logger.info("    [XPM] GST Reconciliation → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] GST Reconciliation failed: %s", exc)
        return None


def _dl_bank_reconciliation(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_bank_reconciliation_report(
            browser=browser,
            client_name=client_name,
            xero_end_date=end_date,
            xero_financial_year=fy,
            xero_start_date=start_date,
            xero_bank_account=Config.get(MC.CFG_XERO_BANK_ACCOUNT, ""),
            window_title=window_title,
            download_directory=str(folder),
            report_file_name=filename,
            xero_report_name=Config.get(
                MC.CFG_XERO_BANK_REC_REPORT_NAME, MC.XERO_RPT_BANK_REC
            ),
            is_no_bank_accounts=False,
            extension=_EXTENSION,
        )
        logger.info("    [XPM] Bank Reconciliation → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Bank Reconciliation failed: %s", exc)
        return None


def _dl_payroll_employee_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_payroll_employee_summary_report(
            browser=browser,
            client_name=client_name,
            xero_end_date=end_date,
            xero_financial_year=fy,
            xero_start_date=start_date,
            window_title=window_title,
            download_directory_path=str(folder),
            xero_report_file_name=filename,
            xero_report_name=Config.get(
                MC.CFG_XERO_PAYROLL_REPORT_NAME, MC.XERO_RPT_PAYROLL
            ),
            extension=_EXTENSION,
        )
        logger.info("    [XPM] Payroll Employee Summary → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Payroll Employee Summary failed: %s", exc)
        return None


def _dl_account_transactions(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_general_ledger_details_report(
            browser=browser,
            client_name=client_name,
            xero_end_date=end_date,
            xero_financial_year=fy,
            xero_start_date=start_date,
            window_title=window_title,
            download_directory=str(folder),
            report_file_name=filename,
            extension=_EXTENSION,
        )
        logger.info("    [XPM] Account Transactions → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Account Transactions failed: %s", exc)
        return None


def _dl_aged_receivables_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
    window_title: str,
) -> Optional[Path]:
    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        xero_blue_download_aged_receivables_summary_report(
            browser=browser,
            client_name=client_name,
            end_date=end_date,
            financial_year=fy,
            is_add_gst_column=Config.get(MC.CFG_XERO_ADD_GST_COLUMN, "false").lower()
            == "true",
            xero_aging_by=Config.get(MC.CFG_XERO_AGING_BY, MC.DEFAULT_AGING_BY),
            window_title=window_title,
            download_directory=str(folder),
            report_file_name=filename,
            extension=_EXTENSION,
        )
        logger.info("    [XPM] Aged Receivables Summary → %s", output_path)
        return output_path
    except Exception as exc:
        logger.warning("    [XPM] Aged Receivables Summary failed: %s", exc)
        return None


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def run_xpm_steps(
    member: Dict[str, Any],
    group_members: List[Dict[str, Any]],
    group_data: Dict[str, Any],
    sm,
) -> Dict[str, Any]:
    """Download all Xero reports for one family member and save into the member output folder.

    Called by the consumer when PMS = "XPM".

    Reports downloaded:
        1.  Balance Sheet                  (TODO — not yet in iaa_rpa_xero_blue)
        2.  Profit and Loss                (TODO — not yet in iaa_rpa_xero_blue)
        3.  Balance Sheet + P&L merged    (pending 1 & 2)
        4.  Aged Payables Summary
        5.  GST Reconciliation
        6.  Bank Reconciliation
        7.  Payroll Employee Summary
        8.  Account Transactions
        9.  Aged Receivables Summary

    Args:
        member:        Combined member record (Client Code, Client Name, Order, etc.).
        group_members: Full sorted combined list (needed to build the folder path).
        group_data:    Queue payload (abn, financialYear, pms).
        sm:            State machine instance (provides browser via initialised_apps).

    Returns:
        Dict with keys: folder_path, downloaded, skipped, status, comment.
    """
    client_code = str(member.get("Client Code", "") or "").strip()
    client_name = str(member.get("Client Name", "") or "").strip()
    fy = str(group_data.get("financialYear", "") or "").strip()
    browser = (getattr(sm, "initialised_apps", {}) or {}).get("browser")
    window_title = Config.get(MC.CFG_XERO_WINDOW_TITLE, MC.DEFAULT_XERO_WINDOW_TITLE)

    start_date, end_date = financial_year_dates(fy)
    logger.info(
        "    [XPM] Starting report downloads for %s (%s)  FY=%s  (%s → %s)",
        client_code,
        client_name,
        fy,
        start_date,
        end_date,
    )

    result: Dict[str, Any] = {
        "folder_path": "",
        "downloaded": [],  # report names successfully downloaded
        "skipped": [],  # report names skipped or failed
        "report_paths": {},  # {report_name: str(file_path)} for each downloaded report
        "status": "success",
        "comment": "",
    }

    try:
        folder = build_member_folder_path(member, group_members, group_data)
        ensure_folder(folder)
        logger.info("    [XPM] Member folder: %s", folder)
        result["folder_path"] = str(folder)
    except Exception as exc:
        result.update(status="failed", comment=f"Folder creation failed: {exc}")
        logger.error("    [XPM] Folder creation failed: %s", exc)
        return result

    if not browser:
        result.update(
            status="failed", comment="Browser not initialised — cannot download reports"
        )
        return result

    # fn: resolves filename from config with {fy} substituted — avoids repeating fy in every call.
    def fn(key: str, default: str) -> str:
        return _resolve_report_filename(key, default, fy)

    # _track: records the downloaded file path on success, or the report name in skipped on failure.
    def _track(path: Optional[Path], name: str) -> None:
        if path is not None:
            result["downloaded"].append(name)
            result["report_paths"][name] = str(path)
        else:
            result["skipped"].append(name)

    bs_path = _dl_balance_sheet(
        browser,
        client_name,
        fy,
        folder,
        fn(MC.CFG_REPORT_BALANCE_SHEET, MC.DEFAULT_BALANCE_SHEET),
    )
    _track(bs_path, MC.REPORT_BALANCE_SHEET)

    pl_path = _dl_profit_and_loss(
        browser,
        client_name,
        fy,
        folder,
        fn(MC.CFG_REPORT_PROFIT_LOSS, MC.DEFAULT_PROFIT_LOSS),
    )
    _track(pl_path, MC.REPORT_PROFIT_LOSS)

    _track(
        _merge_bs_and_pl(
            bs_path,
            pl_path,
            folder,
            fn(MC.CFG_REPORT_MERGED_FINANCIALS, MC.DEFAULT_MERGED_FINANCIALS),
        ),
        MC.REPORT_MERGED_FINANCIALS,
    )
    _track(
        _dl_aged_payables_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_AGED_PAYABLES, MC.DEFAULT_AGED_PAYABLES),
            window_title,
        ),
        MC.REPORT_AGED_PAYABLES,
    )
    _track(
        _dl_gst_reconciliation(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_GST_RECON, MC.DEFAULT_GST_RECON),
            window_title,
        ),
        MC.REPORT_GST_RECON,
    )
    _track(
        _dl_bank_reconciliation(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_BANK_REC, MC.DEFAULT_BANK_REC),
            window_title,
        ),
        MC.REPORT_BANK_REC,
    )
    _track(
        _dl_payroll_employee_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_PAYROLL_EMPLOYEE, MC.DEFAULT_PAYROLL_EMPLOYEE),
            window_title,
        ),
        MC.REPORT_PAYROLL_EMPLOYEE,
    )
    _track(
        _dl_account_transactions(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_ACCOUNT_TRANS, MC.DEFAULT_ACCOUNT_TRANS),
            window_title,
        ),
        MC.REPORT_ACCOUNT_TRANS,
    )
    _track(
        _dl_aged_receivables_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_AGED_RECEIVABLES, MC.DEFAULT_AGED_RECEIVABLES),
            window_title,
        ),
        MC.REPORT_AGED_RECEIVABLES,
    )

    n_ok, n_skip = len(result["downloaded"]), len(result["skipped"])
    result["status"] = (
        "success" if n_skip == 0 else ("partial" if n_ok > 0 else "failed")
    )
    result["comment"] = (
        f"Reports: {n_ok} downloaded, {n_skip} skipped — folder: {folder.name}"
    )
    logger.info("    [XPM] Reports complete — %s downloaded, %s skipped", n_ok, n_skip)
    for name, path in result["report_paths"].items():
        logger.info("    [XPM]   %-35s → %s", name, path)
    if result["skipped"]:
        logger.info("    [XPM] Skipped: %s", result["skipped"])
    return result

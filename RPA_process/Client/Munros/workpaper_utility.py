"""
Workpaper Utility — Munros Client
====================================
Shared utilities used across all Munros process modules:

    Excel reading / grouping / sorting   — used by producer + consumer
    Member folder path building          — used by xero_processes, myob_processes, ato_processes

Keeping path utilities here avoids duplication: every report-download module
imports build_member_folder_path / ensure_folder / report_filename from one place.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from dateutil.parser import parse as parse_date
from dateutil.parser import ParserError

from iaa_rpa_framework import Config, RPASystemException
from RPA_process.Client.Munros import munros_constants as MC

logger = logging.getLogger("IARPA." + __name__)

# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------


def read_excel(
    xlsx_path: Path, sheet_name: Optional[str] = None
) -> List[Dict[str, Any]]:
    """Read all non-blank rows from an Excel workbook as a list of dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    logger.info("Reading sheet '%s' from %s", ws.title, xlsx_path)

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise RPASystemException(f"Excel sheet '{ws.title}' is empty: {xlsx_path}")

    headers = [
        str(h).strip() if h is not None else f"_col{i}" for i, h in enumerate(rows[0])
    ]
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        records.append(
            {
                headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                for i in range(len(headers))
            }
        )

    logger.info("Loaded %s record(s) from %s", len(records), xlsx_path)
    return records


def validate_headers(
    records: List[Dict[str, Any]], required: List[str], label: str
) -> None:
    """Raise RPASystemException if any required column is missing from the first record."""
    if not records:
        return
    present = set(records[0].keys())
    missing = [col for col in required if col not in present]
    if missing:
        raise RPASystemException(
            f"{label}: missing required column(s): {', '.join(missing)}. "
            f"Found: {', '.join(sorted(present))}"
        )


# ---------------------------------------------------------------------------
# ABN normalisation
# ---------------------------------------------------------------------------


def normalise_abn(raw: Any) -> str:
    """Return digits-only ABN string for reliable comparison."""
    return re.sub(r"\D", "", str(raw or ""))


def build_abn_index(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Index Excel 2 records by normalised ABN."""
    index: Dict[str, List[Dict[str, Any]]] = {}
    for row in records:
        key = normalise_abn(row.get("ABN", ""))
        if key:
            index.setdefault(key, []).append(row)
    return index


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------


def determine_parent_code(matched_row: Dict[str, Any]) -> str:
    """Return the group parent code for a row matched from Excel 2.

    If Client Parent Code is blank or equals Client Code, the row itself
    is the parent — return Client Code.  Otherwise return Client Parent Code.
    """
    client_code = str(matched_row.get("Client Code", "") or "").strip()
    client_parent_code = str(matched_row.get("Client Parent Code", "") or "").strip()

    if not client_parent_code or client_parent_code == client_code:
        return client_code
    return client_parent_code


def is_parent_record(row: Dict[str, Any]) -> bool:
    """Return True if this Excel 2 row is the parent of its group."""
    client_code = str(row.get("Client Code", "") or "").strip()
    client_parent_code = str(row.get("Client Parent Code", "") or "").strip()
    return not client_parent_code or client_code == client_parent_code


def collect_group_members(
    details_records: List[Dict[str, Any]], parent_code: str
) -> List[Dict[str, Any]]:
    """Return all Excel 2 rows belonging to the family group identified by parent_code.

    Includes:
        - The parent itself  (Client Code == parent_code  AND  is_parent_record)
        - All children       (Client Parent Code == parent_code)
    """
    members: List[Dict[str, Any]] = []
    for row in details_records:
        client_code = str(row.get("Client Code", "") or "").strip()
        client_parent_code = str(row.get("Client Parent Code", "") or "").strip()

        is_self_parent = client_code == parent_code and is_parent_record(row)
        is_child = client_parent_code == parent_code

        if is_self_parent or is_child:
            members.append(row)

    return members


# ---------------------------------------------------------------------------
# Sorting
# ---------------------------------------------------------------------------


def _parse_dob(dob_str: str) -> Optional[float]:
    """Parse a date-of-birth string and return a Unix timestamp, or None on failure."""
    dob_str = dob_str.strip()
    if not dob_str:
        return None
    try:
        return parse_date(dob_str, dayfirst=True).timestamp()
    except (ParserError, ValueError, OverflowError):
        logger.debug("Could not parse DOB: '%s'", dob_str)
        return None


def age_sort_key(row: Dict[str, Any]) -> Tuple:
    """Secondary sort key: oldest first, records without DOB go last."""
    dob_ts = _parse_dob(str(row.get("Date Of Birth", "") or ""))
    has_dob = 0 if dob_ts is not None else 1
    return (has_dob, dob_ts if dob_ts is not None else float("inf"))


def sort_members(members: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort group members with the Munros ordering rule:

    1. The single oldest male is always first (Order = 1).
    2. All remaining members are sorted oldest → youngest regardless of sex.
       Records without a Date Of Birth are placed last.
    """
    is_male = lambda r: str(r.get("Sex", "") or "").strip().upper() in ("M", "MALE")

    males = [r for r in members if is_male(r)]
    others = [r for r in members if not is_male(r)]

    if males:
        males_with_dob = [
            (r, _parse_dob(str(r.get("Date Of Birth", "") or ""))) for r in males
        ]
        valid = [(r, ts) for r, ts in males_with_dob if ts is not None]
        oldest_male = min(valid, key=lambda x: x[1])[0] if valid else males[0]
        remaining = [r for r in males if r is not oldest_male] + others
    else:
        oldest_male = None
        remaining = list(others)

    result = []
    if oldest_male is not None:
        result.append(oldest_male)
    result.extend(sorted(remaining, key=age_sort_key))
    return result


# ---------------------------------------------------------------------------
# Client name derivation
# ---------------------------------------------------------------------------


def build_client_name(row: Dict[str, Any]) -> str:
    """Derive display name from Excel 2 First Name + Last Name.

    If First Name is blank (company, trust, etc.) return Last Name only.
    """
    first = str(row.get("First Name", "") or "").strip()
    last = str(row.get("Last Name", "") or "").strip()
    return f"{first} {last}".strip() if first else last


# ---------------------------------------------------------------------------
# Combined record builder
# ---------------------------------------------------------------------------


def build_combined_records(
    group_abn: str,
    financial_year: str,
    excel1_client_name: str,
    group_members: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Sort group members and return a list of combined payload dicts.

    Each record contains the required output columns plus extra fields
    the consumer will use (TFN, DOB, Sex, Client Partner, is_parent).

    Args:
        group_abn:          Normalised ABN string from Excel 1.
        financial_year:     Financial Year from Excel 1.
        excel1_client_name: Client Name from Excel 1 (kept as reference).
        group_members:      All Excel 2 rows in the family group (unsorted).

    Returns:
        List of combined dicts, one per family member, with Order assigned.
    """
    sorted_members = sort_members(group_members)

    combined: List[Dict[str, Any]] = []
    for order, member in enumerate(sorted_members, start=1):
        combined.append(
            {
                # Required output columns
                "Client Name": build_client_name(member),
                "ABN": group_abn,
                "Financial Year": financial_year,
                "Client Code": str(member.get("Client Code", "") or "").strip(),
                "Client Parent Code": str(
                    member.get("Client Parent Code", "") or ""
                ).strip(),
                "Order": order,
                # Consumer fields
                "TFN": str(member.get("TFN", "") or "").strip(),
                "Date Of Birth": str(member.get("Date Of Birth", "") or "").strip(),
                "Sex": str(member.get("Sex", "") or "").strip(),
                "Client Partner": str(member.get("Client Partner", "") or "").strip(),
                "is_parent": is_parent_record(member),
                # Reference
                "excel1_client_name": excel1_client_name,
            }
        )

    return combined


# ---------------------------------------------------------------------------
# Member output folder utilities  (shared by xero_reports, myob_processes, ato_processes)
# ---------------------------------------------------------------------------


def sanitize_path_part(value: str) -> str:
    """Remove characters that are illegal in Windows folder/file names."""
    return re.sub(r'[<>:"/\\|?*]', "_", str(value or "").strip())


def order_to_letter(order: int) -> str:
    """Convert a 1-based Order number to a capital letter: 1→A, 2→B, …, 26→Z."""
    if 1 <= order <= 26:
        return chr(64 + order)
    return str(order)


def build_member_folder_path(
    member: Dict[str, Any],
    group_members: List[Dict[str, Any]],
    group_data: Dict[str, Any],
) -> Path:
    """Build and return the output folder Path for one family member.

    Structure:
        {Munros_ReportRootPath}
            \\ {parent_code} {parent_name}
                \\ {financial_year}
                    \\ Tax
                        \\ {order_letter} {member_name}

    Args:
        member:        The member being processed.
        group_members: Full sorted combined list (used to find the parent member).
        group_data:    Queue payload (provides financialYear).

    Raises:
        RPASystemException: When Munros_ReportRootPath config key is not set.
    """
    root = Config.get(MC.CFG_REPORT_ROOT_PATH, "")
    if not root:
        raise RPASystemException(
            f"Config key '{MC.CFG_REPORT_ROOT_PATH}' is not set. "
            "Cannot build member output folder path."
        )

    parent = next((m for m in group_members if m.get("is_parent")), group_members[0])
    level1 = sanitize_path_part(
        f"{parent.get('Client Code', '')} {parent.get('Client Name', '')}".strip()
    )
    level2 = sanitize_path_part(str(group_data.get("financialYear", "") or "").strip())
    order = int(member.get("Order", 1))
    level4 = sanitize_path_part(
        f"{order_to_letter(order)} {member.get('Client Name', '')}".strip()
    )

    return Path(root) / level1 / level2 / "Tax" / level4


def ensure_folder(path: Path) -> None:
    """Create the folder (and all parents) if it does not already exist."""
    path.mkdir(parents=True, exist_ok=True)
    logger.info("    Folder ready: %s", path)


def report_filename(config_key: str, default_template: str, fy: str) -> str:
    """Resolve a report filename from config, substituting {fy} with the financial year."""
    template = Config.get(config_key, default_template) or default_template
    return template.replace("{fy}", fy).strip()


def financial_year_dates(financial_year: str) -> tuple:
    """Return (start_date, end_date) strings for an Australian financial year.

    Australian FY runs 1 July to 30 June.
        FY 2025  →  start = "1 Jul 2024",  end = "30 Jun 2025"

    Used by xero_processes, myob_processes, and ato_processes when passing
    explicit date ranges to report download functions.

    Args:
        financial_year: Year string, e.g. "2025".

    Returns:
        Tuple of (start_date, end_date) in "d MMM YYYY" format accepted by
        iaa_rpa_xero_blue report functions.  Returns ("", "") on parse failure
        so callers can still pass the empty strings as a default fallback.
    """
    try:
        fy = int(str(financial_year).strip())
    except (ValueError, TypeError):
        logger.warning(
            "financial_year_dates: cannot parse financial year '%s' — returning empty dates",
            financial_year,
        )
        return ("", "")

    return (f"1 Jul {fy - 1}", f"30 Jun {fy}")

"""
MYOB Processes — Munros Client
================================
Single public entry point: run_myob_steps()

Called by the consumer for each family member when PMS = "MYOB".

Covers two responsibilities:
    1. Report downloads  — all MYOB reports saved to the member output folder
                           (placeholder _dl_* functions; implement as MYOB API is confirmed)
    2. MYOB Web steps    — session check, client selection, document upload, email draft
                           (placeholder helpers; implement as process map is confirmed)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from iaa_rpa_framework import Config

from RPA_process.Client.Munros.workpaper_utility import (
    build_member_folder_path,
    ensure_folder,
    report_filename as _resolve_report_filename,
    financial_year_dates,
)
from RPA_process.Client.Munros import munros_constants as MC

logger = logging.getLogger("IARPA." + __name__)

# Local alias so the extension string isn't repeated throughout every _dl_* function.
_EXTENSION = MC.REPORT_EXTENSION


# ===========================================================================
# Section A — Report download helpers
# (placeholder — implement as MYOB API/web functions become available)
# ===========================================================================


def _dl_balance_sheet(
    browser,
    client_name: str,
    fy: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement once MYOB Balance Sheet report download is available."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Balance Sheet — TODO: implement (client=%s, fy=%s, file=%s)",
        client_name,
        fy,
        output_path.name,
    )
    _ = browser  # will be used once the download function is available
    return None


def _dl_profit_and_loss(
    browser,
    client_name: str,
    fy: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement once MYOB Profit and Loss report download is available."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Profit and Loss — TODO: implement (client=%s, fy=%s, file=%s)",
        client_name,
        fy,
        output_path.name,
    )
    _ = browser  # will be used once the download function is available
    return None


def _merge_bs_and_pl(
    bs_path: Optional[Path],
    pl_path: Optional[Path],
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """Merge Balance Sheet and P&L into one workbook with two named sheets.

    Runs automatically once both source files are downloaded.
    Sheet names are fixed by the workpaper standard — not config-driven.
    """
    if not bs_path or not pl_path:
        logger.info(
            "    [MYOB] Merged financials skipped — BS and/or P&L not yet downloaded (TODO)"
        )
        return None

    output_path = folder / f"{filename}{_EXTENSION}"
    try:
        merged_wb = openpyxl.Workbook()
        merged_wb.remove(merged_wb.active)
        for source_path, sheet_title in (
            (bs_path, "Balance Sheet"),
            (pl_path, "Profit and Loss"),
        ):
            source_wb = openpyxl.load_workbook(source_path, data_only=True)
            dest_ws = merged_wb.create_sheet(title=sheet_title)
            for row in source_wb.active.iter_rows(values_only=True):
                dest_ws.append(list(row))
            source_wb.close()
        merged_wb.save(output_path)
        merged_wb.close()
        logger.info("    [MYOB] Merged financials saved: %s", output_path.name)
        return output_path
    except Exception as exc:
        logger.warning("    [MYOB] Merged financials failed: %s", exc)
        return None


def _dl_aged_payables_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB Aged Payables Summary download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Aged Payables Summary — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


def _dl_gst_reconciliation(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB GST Reconciliation download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] GST Reconciliation — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


def _dl_bank_reconciliation(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB Bank Reconciliation download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Bank Reconciliation — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


def _dl_payroll_employee_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB Payroll Employee Summary download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Payroll Employee Summary — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


def _dl_account_transactions(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB Account Transactions download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Account Transactions — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


def _dl_aged_receivables_summary(
    browser,
    client_name: str,
    fy: str,
    start_date: str,
    end_date: str,
    folder: Path,
    filename: str,
) -> Optional[Path]:
    """TODO: implement MYOB Aged Receivables Summary download."""
    output_path = folder / f"{filename}{_EXTENSION}"
    logger.info(
        "    [MYOB] Aged Receivables Summary — TODO: implement (client=%s, fy=%s, %s→%s, file=%s)",
        client_name,
        fy,
        start_date,
        end_date,
        output_path.name,
    )
    _ = browser
    return None


# ===========================================================================
# Public entry point
# ===========================================================================


def run_myob_steps(
    member: Dict[str, Any],
    group_members: List[Dict[str, Any]],
    group_data: Dict[str, Any],
    sm,
) -> Dict[str, Any]:
    """Download all MYOB reports and run MYOB Web steps for one family member.

    Called by the consumer when PMS = "MYOB".

    Reports downloaded (all TODO — placeholders in place):
        1.  Balance Sheet
        2.  Profit and Loss
        3.  Balance Sheet + P&L merged
        4.  Aged Payables Summary
        5.  GST Reconciliation
        6.  Bank Reconciliation
        7.  Payroll Employee Summary
        8.  Account Transactions
        9.  Aged Receivables Summary

    Args:
        member:        Combined member record (Client Code, Client Name, TFN, etc.).
        group_members: Full sorted combined list (needed to build the folder path).
        group_data:    Queue payload (abn, financialYear, pms).
        sm:            State machine instance (browser, myob_connection, config).

    Returns:
        Dict with keys: folder_path, downloaded, skipped, report_paths, status, comment.
    """
    client_code = str(member.get("Client Code", "") or "").strip()
    client_name = str(member.get("Client Name", "") or "").strip()
    fy = str(group_data.get("financialYear", "") or "").strip()
    browser = (getattr(sm, "initialised_apps", {}) or {}).get("browser")

    start_date, end_date = financial_year_dates(fy)
    logger.info(
        "    [MYOB] Starting MYOB steps for %s (%s)  FY=%s  (%s → %s)",
        client_code,
        client_name,
        fy,
        start_date,
        end_date,
    )

    result: Dict[str, Any] = {
        "folder_path": "",
        "downloaded": [],  # report names successfully downloaded
        "skipped": [],  # report names skipped or failed
        "report_paths": {},  # {report_name: str(file_path)} for each downloaded report
        "status": "success",
        "comment": "",
    }

    # Build and create the member output folder
    try:
        folder = build_member_folder_path(member, group_members, group_data)
        ensure_folder(folder)
        logger.info("    [MYOB] Member folder: %s", folder)
        result["folder_path"] = str(folder)
    except Exception as exc:
        result.update(status="failed", comment=f"Folder creation failed: {exc}")
        logger.error("    [MYOB] Folder creation failed: %s", exc)
        return result

    # fn: resolves filename from config with {fy} substituted — avoids repeating fy in every call.
    def fn(key: str, default: str) -> str:
        return _resolve_report_filename(key, default, fy)

    # _track: records the downloaded file path on success, or the report name in skipped on failure.
    def _track(path: Optional[Path], name: str) -> None:
        if path is not None:
            result["downloaded"].append(name)
            result["report_paths"][name] = str(path)
        else:
            result["skipped"].append(name)

    # --- Section A: report downloads ---
    bs_path = _dl_balance_sheet(
        browser,
        client_name,
        fy,
        folder,
        fn(MC.CFG_REPORT_BALANCE_SHEET, MC.DEFAULT_BALANCE_SHEET),
    )
    _track(bs_path, MC.REPORT_BALANCE_SHEET)

    pl_path = _dl_profit_and_loss(
        browser,
        client_name,
        fy,
        folder,
        fn(MC.CFG_REPORT_PROFIT_LOSS, MC.DEFAULT_PROFIT_LOSS),
    )
    _track(pl_path, MC.REPORT_PROFIT_LOSS)

    _track(
        _merge_bs_and_pl(
            bs_path,
            pl_path,
            folder,
            fn(MC.CFG_REPORT_MERGED_FINANCIALS, MC.DEFAULT_MERGED_FINANCIALS),
        ),
        MC.REPORT_MERGED_FINANCIALS,
    )
    _track(
        _dl_aged_payables_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_AGED_PAYABLES, MC.DEFAULT_AGED_PAYABLES),
        ),
        MC.REPORT_AGED_PAYABLES,
    )
    _track(
        _dl_gst_reconciliation(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_GST_RECON, MC.DEFAULT_GST_RECON),
        ),
        MC.REPORT_GST_RECON,
    )
    _track(
        _dl_bank_reconciliation(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_BANK_REC, MC.DEFAULT_BANK_REC),
        ),
        MC.REPORT_BANK_REC,
    )
    _track(
        _dl_payroll_employee_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_PAYROLL_EMPLOYEE, MC.DEFAULT_PAYROLL_EMPLOYEE),
        ),
        MC.REPORT_PAYROLL_EMPLOYEE,
    )
    _track(
        _dl_account_transactions(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_ACCOUNT_TRANS, MC.DEFAULT_ACCOUNT_TRANS),
        ),
        MC.REPORT_ACCOUNT_TRANS,
    )
    _track(
        _dl_aged_receivables_summary(
            browser,
            client_name,
            fy,
            start_date,
            end_date,
            folder,
            fn(MC.CFG_REPORT_AGED_RECEIVABLES, MC.DEFAULT_AGED_RECEIVABLES),
        ),
        MC.REPORT_AGED_RECEIVABLES,
    )

    n_ok, n_skip = len(result["downloaded"]), len(result["skipped"])
    result["status"] = (
        "success" if n_skip == 0 else ("partial" if n_ok > 0 else "failed")
    )
    result["comment"] = (
        f"Reports: {n_ok} downloaded, {n_skip} skipped — folder: {folder.name}"
    )
    logger.info("    [MYOB] Reports complete — %s downloaded, %s skipped", n_ok, n_skip)
    if result["skipped"]:
        logger.info("    [MYOB] Skipped: %s", result["skipped"])

    logger.info("    [MYOB] MYOB steps complete for %s", client_code)
    return result
